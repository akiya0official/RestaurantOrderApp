# income_checker.py
from openpyxl import load_workbook
from datetime import datetime

file_name = "restaurant_orders.xlsx"
today = datetime.now().strftime("%Y-%m-%d")
wb = load_workbook(file_name)
ws = wb.active

total_income = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] == today:
        total_income += row[4]

print(f"අද දිනයේ මුළු ආදායම: Rs.{total_income}")