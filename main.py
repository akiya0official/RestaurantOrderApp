# main.py
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

items = {
    "Sotis": 100,
    "Rice": 250,
    "Juice": 150
}

file_name = "restaurant_orders.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Item", "Qty", "Unit Price", "Total"])
    wb.save(file_name)

root = tk.Tk()
root.title("Restaurant Order System")

tk.Label(root, text="Item Select කරන්න").grid(row=0, column=0)
item_var = tk.StringVar(root)
item_var.set("Sotis")
tk.OptionMenu(root, item_var, *items.keys()).grid(row=0, column=1)

tk.Label(root, text="Qty ගන්න").grid(row=1, column=0)
qty_entry = tk.Entry(root)
qty_entry.grid(row=1, column=1)

def save_order():
    item = item_var.get()
    try:
        qty = int(qty_entry.get())
        price = items[item]
        total = qty * price
        date = datetime.now().strftime("%Y-%m-%d")

        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([date, item, qty, price, total])
        wb.save(file_name)

        messagebox.showinfo("Saved", f"{item} {qty}ක් = Rs.{total} saved!")
        qty_entry.delete(0, tk.END)
    except:
        messagebox.showerror("Error", "Qty එක integer එකක් විය යුතුයි")

tk.Button(root, text="Save Order", command=save_order).grid(row=2, column=0, columnspan=2)

root.mainloop()